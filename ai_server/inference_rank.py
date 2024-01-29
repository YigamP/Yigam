import json
import numpy as np 
import pandas as pd
from pytorch_metric_learning import miners, losses, distances
import os 
from transformers import * 
import torch 
import torch.nn as nn
import torch.nn.functional as F 
from torch.utils.data import Dataset, TensorDataset, DataLoader, RandomSampler, SequentialSampler 
import pickle 
import time 
from sklearn.model_selection import train_test_split, KFold
from tqdm.auto import tqdm 
import random 
import math 
import faiss
from tqdm import tqdm
from openpyxl import load_workbook

def get_data(input_query):

    def read_jsonl_file(file_path):
        arr = [] 
        with open(file_path, 'r') as file:
            for line in file:
                json_data = json.loads(line)
                arr.append(json_data)
        return arr
    
    def read_excel_to_list_of_lists(filename):
        # Load the workbook and select the active worksheet
        workbook = load_workbook(filename)
        sheet = workbook.active
        
        # Read the rows into a list of lists
        data = []
        idx = 0
        for row in sheet.iter_rows(values_only=True):
            idx += 1
            if idx == 1:
                continue

            temp = {
                "source": row[0],
                "book": row[1],
                "paragraph": row[2],
                "question": row[3].replace('\n\n', ' ').replace('\n', ' ') if row[3] != None else '' if row[3] == None else row[3],
                "answer": row[4].replace('\n\n', ' ').replace('\n', ' ').replace('도움이 됐길 바랍니다.', '').replace('감사합니다.', '') if row[4] != None else '' if row[4] == None else row[4],
            }
            data.append(temp)
        
        return data
    
    import glob 
    print("!!!====== files   ", glob.glob("./*.xlsx"))
    filename = 'answer.xlsx'
    filename = 'answer.xlsx'
    excel_data = read_excel_to_list_of_lists(filename)


    sources, books, paragraphs, questions, answers = [], [], [], [], []
    for i in range(len(excel_data)):
        sources.append(excel_data[i]["source"])
        books.append(excel_data[i]["book"])
        paragraphs.append(excel_data[i]["paragraph"])
        questions.append(excel_data[i]["question"])
        answers.append(excel_data[i]["answer"])


    model_name = "monologg/kobigbird-bert-base" 
    q_tokenizer = AutoTokenizer.from_pretrained(model_name) 
    p_tokenizer = AutoTokenizer.from_pretrained(model_name) 

    query_index_dict = {} 
    for i in range(len(questions)): 
        query_index_dict[questions[i]] = [] 
    for i in range(len(questions)): 
        query_index_dict[questions[i]].append(i) # query string: list of relevant candidate ids 


    device = torch.device("cuda" if torch.cuda.is_available() else "cpu") 

    # define model structure
    class MeanPooling(nn.Module): 
        def __init__(self): 
            super(MeanPooling, self).__init__() 
        def forward(self, last_hidden_state, attention_masks): 
            input_mask_expanded = attention_masks.unsqueeze(-1).expand(last_hidden_state.size()).float() 
            sum_embeddings = torch.sum(last_hidden_state * input_mask_expanded, 1) 
            sum_mask = input_mask_expanded.sum(1) 
            sum_mask = torch.clamp(sum_mask, min=1e-9) 
            mean_embeddings = sum_embeddings / sum_mask 
            return mean_embeddings

    class BertEmbedder(nn.Module): 
        def __init__(self, plm):
            super(BertEmbedder, self).__init__() 
            self.encoder = AutoModel.from_pretrained(plm)
            self.mean_pooler = MeanPooling() 
        def forward(self, input_ids, attn_masks): 
            x = self.encoder(input_ids, attn_masks)[0] 
            x = self.mean_pooler(x, attn_masks) 
            return x 

    q_encoder = BertEmbedder(plm=model_name) 
    # q_checkpoint = torch.load("KR_lawqa_KoBigBird_query_encoder.pt", map_location=torch.device('cpu'))
    q_checkpoint = torch.load("KoBigBird_query_encoder.pt", map_location=torch.device('cpu'))
    q_encoder.load_state_dict(q_checkpoint, strict=False) 
    q_encoder.to(device) 
    q_encoder.eval() 

    p_encoder = BertEmbedder(plm=model_name) 
    # p_checkpoint = torch.load("KR_lawqa_KoBigBird_passage_encoder.pt", map_location=torch.device('cpu'))
    p_checkpoint = torch.load("KoBigBird_passage_encoder.pt", map_location=torch.device('cpu'))
    p_encoder.load_state_dict(p_checkpoint, strict=False) 
    p_encoder.to(device) 
    p_encoder.eval()  

    test_recall_1, test_recall_5 = [], [] # focus on R@1, R@5 

    # read saved index 
    index = faiss.read_index("candidate_embeddings.index")

    def get_ranks(query_text, topK=100):
        with torch.no_grad(): 
            encoded_query = q_tokenizer(str(query_text), max_length=512, truncation=True, padding="max_length", return_tensors="pt").to(device)
            input_id = encoded_query["input_ids"] 
            attn_mask = encoded_query["attention_mask"] 
            q_emb = q_encoder(input_id, attn_mask) 
            q_emb = q_emb.detach().cpu().numpy() 
            distances, indices = index.search(q_emb, 1000) 
            topK_texts = [] 
            topK_dist = []
            topK_indices = []
            topK_questions = []
            topK_answers = []
            for i in range(topK):
                try:
                    # topK_texts.append([questions[indices[0][i]], answers[indices[0][i]] ])
                    # print("!!!========= topK   question and answer   ", [questions[indices[0][i]], answers[indices[0][i]] ])
                    topK_questions.append(questions[indices[0][i]])
                    topK_answers.append(answers[indices[0][i]])

                    topK_dist.append(distances[0][i])
                    topK_indices.append(int(indices[0][i]))

                    # title = courts[indices[0][i]] + " " + f"{dates[indices[0][i]][:4]}.{dates[indices[0][i]][4:6]}.{dates[indices[0][i]][6:]}." + " " + types[indices[0][i]] + " " + references[indices[0][i]] + " [" + titles[indices[0][i]] + "]"
                    # topK_texts.append(passages[indices[0][i]])
                    # topK_texts.append(title)
                    # topK_dist.append(distances[0][i])
                    # topK_indices.append(int(indices[0][i]))
                except:
                    pass
            topK_dist = [min(x+25.0, 100.0) for x in topK_dist]
            topK_texts = [topK_questions, topK_answers]
        return topK_texts, topK_dist, topK_indices
    
    ranked = get_ranks(input_query, topK=100)

    return ranked

# temp = get_data('(나)의 화자가 거울을 통해 거짓되고 부조리한 인간의 현실을 응시한다는게 거울 안에 보이는 모습을 응시하는걸로 이해했는데 거울 안 보다 시에는 \"거울 앞에서... 뒤집은 현실의 뒤집은 마을의 주민\"이라고 ')

# for i in range(0, len(temp[0])):
#     print("==================================================================  i " , i , "   =============   ", temp[1][i], "    ", temp[2][i])
#     print(json.dumps(temp[0][i],indent=4, ensure_ascii=False))

